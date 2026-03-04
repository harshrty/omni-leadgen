import sqlite3, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

conn = sqlite3.connect('leads.db')
cur = conn.cursor()

# Get actual column names from DB
cur.execute("PRAGMA table_info(leads)")
db_cols = [row[1] for row in cur.fetchall()]
print("DB columns:", db_cols)

# Fetch all leads with decision maker email
cur.execute("SELECT * FROM leads WHERE decision_maker_email IS NOT NULL AND decision_maker_email != ''")
rows = cur.fetchall()
print(f"Exporting {len(rows)} leads with contacts...")

if not rows:
    print("No leads to export!")
    conn.close()
    exit()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Leads"

# Use actual DB columns as headers, prettify them
display_headers = [col.replace("_", " ").title() for col in db_cols]

# Header row styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
for col_idx, h in enumerate(display_headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Data rows
for row_idx, row in enumerate(rows, 2):
    for col_idx, val in enumerate(row, 1):
        ws.cell(row=row_idx, column=col_idx, value=val or "")

# Auto-width columns
for col_idx in range(1, len(db_cols) + 1):
    max_len = len(display_headers[col_idx - 1])
    for row_idx in range(2, min(len(rows) + 2, 50)):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val:
            max_len = max(max_len, min(len(str(val)), 50))
    ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

# Freeze header + auto filter
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save("leads_export.xlsx")
print("Saved: leads_export.xlsx")
conn.close()